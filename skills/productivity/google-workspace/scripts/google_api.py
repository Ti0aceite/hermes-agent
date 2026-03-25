#!/usr/bin/env python3
"""Google Workspace API CLI for Hermes Agent.

A thin CLI wrapper around Google's Python client libraries.
Authenticates using the token stored by setup.py.

Usage:
  python google_api.py gmail search "is:unread" [--max 10]
  python google_api.py gmail get MESSAGE_ID
  python google_api.py gmail send --to user@example.com --subject "Hi" --body "Hello"
  python google_api.py gmail reply MESSAGE_ID --body "Thanks"
  python google_api.py calendar list [--from DATE] [--to DATE] [--calendar primary]
  python google_api.py calendar create --summary "Meeting" --start DATETIME --end DATETIME
  python google_api.py drive search "budget report" [--max 10]
  python google_api.py contacts list [--max 20]
  python google_api.py sheets get SHEET_ID RANGE
  python google_api.py sheets update SHEET_ID RANGE --values '[[...]]'
  python google_api.py sheets append SHEET_ID RANGE --values '[[...]]'
  python google_api.py docs get DOC_ID
"""

import argparse
import base64
import json
import os
import sys
from datetime import datetime, timedelta, timezone
from email.mime.text import MIMEText
from pathlib import Path

HERMES_HOME = Path(os.getenv("HERMES_HOME", Path.home() / ".hermes"))
TOKEN_PATH = HERMES_HOME / "google_token.json"
GMAIL_TOKEN_PATH = HERMES_HOME / "google_gmail_token.json"

SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/gmail.send",
    "https://www.googleapis.com/auth/gmail.modify",
    "https://www.googleapis.com/auth/calendar",
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/contacts.readonly",
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/documents.readonly",
]


def get_credentials(token_path=None):
    """Load and refresh credentials from token file.

    Args:
        token_path: Path to the token JSON file.  Defaults to TOKEN_PATH
                    (google_token.json).  Pass GMAIL_TOKEN_PATH for Gmail.
    """
    path = token_path or TOKEN_PATH
    if not path.exists():
        print(f"Not authenticated ({path.name}). Run the setup script first:", file=sys.stderr)
        if path == GMAIL_TOKEN_PATH:
            print(f"  python {Path(__file__).parent / 'setup.py'} --auth-url --token-name gmail", file=sys.stderr)
        else:
            print(f"  python {Path(__file__).parent / 'setup.py'}", file=sys.stderr)
        sys.exit(1)

    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request

    creds = Credentials.from_authorized_user_file(str(path))
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
        path.write_text(creds.to_json())
    if not creds.valid:
        print(f"Token is invalid ({path.name}). Re-run setup.", file=sys.stderr)
        sys.exit(1)
    return creds


def build_service(api, version, token_path=None):
    """Build a Google API service client.

    Args:
        api: API name (e.g. "gmail", "drive").
        version: API version (e.g. "v1", "v3").
        token_path: Optional path to the token file.  Defaults to TOKEN_PATH.
    """
    from googleapiclient.discovery import build
    return build(api, version, credentials=get_credentials(token_path))


# =========================================================================
# Gmail
# =========================================================================

def gmail_search(args):
    service = build_service("gmail", "v1", GMAIL_TOKEN_PATH)
    results = service.users().messages().list(
        userId="me", q=args.query, maxResults=args.max
    ).execute()
    messages = results.get("messages", [])
    if not messages:
        print("No messages found.")
        return

    output = []
    for msg_meta in messages:
        msg = service.users().messages().get(
            userId="me", id=msg_meta["id"], format="metadata",
            metadataHeaders=["From", "To", "Subject", "Date"],
        ).execute()
        headers = {h["name"]: h["value"] for h in msg.get("payload", {}).get("headers", [])}
        output.append({
            "id": msg["id"],
            "threadId": msg["threadId"],
            "from": headers.get("From", ""),
            "to": headers.get("To", ""),
            "subject": headers.get("Subject", ""),
            "date": headers.get("Date", ""),
            "snippet": msg.get("snippet", ""),
            "labels": msg.get("labelIds", []),
        })
    print(json.dumps(output, indent=2, ensure_ascii=False))


def gmail_get(args):
    service = build_service("gmail", "v1", GMAIL_TOKEN_PATH)
    msg = service.users().messages().get(
        userId="me", id=args.message_id, format="full"
    ).execute()

    headers = {h["name"]: h["value"] for h in msg.get("payload", {}).get("headers", [])}

    # Extract body text
    body = ""
    payload = msg.get("payload", {})
    if payload.get("body", {}).get("data"):
        body = base64.urlsafe_b64decode(payload["body"]["data"]).decode("utf-8", errors="replace")
    elif payload.get("parts"):
        for part in payload["parts"]:
            if part.get("mimeType") == "text/plain" and part.get("body", {}).get("data"):
                body = base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8", errors="replace")
                break
        if not body:
            for part in payload["parts"]:
                if part.get("mimeType") == "text/html" and part.get("body", {}).get("data"):
                    body = base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8", errors="replace")
                    break

    result = {
        "id": msg["id"],
        "threadId": msg["threadId"],
        "from": headers.get("From", ""),
        "to": headers.get("To", ""),
        "subject": headers.get("Subject", ""),
        "date": headers.get("Date", ""),
        "labels": msg.get("labelIds", []),
        "body": body,
    }
    print(json.dumps(result, indent=2, ensure_ascii=False))


def gmail_send(args):
    service = build_service("gmail", "v1", GMAIL_TOKEN_PATH)
    message = MIMEText(args.body, "html" if args.html else "plain")
    message["to"] = args.to
    message["subject"] = args.subject
    if args.cc:
        message["cc"] = args.cc

    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
    body = {"raw": raw}

    if args.thread_id:
        body["threadId"] = args.thread_id

    result = service.users().messages().send(userId="me", body=body).execute()
    print(json.dumps({"status": "sent", "id": result["id"], "threadId": result.get("threadId", "")}, indent=2))


def gmail_reply(args):
    service = build_service("gmail", "v1", GMAIL_TOKEN_PATH)
    # Fetch original to get thread ID and headers
    original = service.users().messages().get(
        userId="me", id=args.message_id, format="metadata",
        metadataHeaders=["From", "Subject", "Message-ID"],
    ).execute()
    headers = {h["name"]: h["value"] for h in original.get("payload", {}).get("headers", [])}

    subject = headers.get("Subject", "")
    if not subject.startswith("Re:"):
        subject = f"Re: {subject}"

    message = MIMEText(args.body)
    message["to"] = headers.get("From", "")
    message["subject"] = subject
    if headers.get("Message-ID"):
        message["In-Reply-To"] = headers["Message-ID"]
        message["References"] = headers["Message-ID"]

    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
    body = {"raw": raw, "threadId": original["threadId"]}

    result = service.users().messages().send(userId="me", body=body).execute()
    print(json.dumps({"status": "sent", "id": result["id"], "threadId": result.get("threadId", "")}, indent=2))


def gmail_labels(args):
    service = build_service("gmail", "v1", GMAIL_TOKEN_PATH)
    results = service.users().labels().list(userId="me").execute()
    labels = [{"id": l["id"], "name": l["name"], "type": l.get("type", "")} for l in results.get("labels", [])]
    print(json.dumps(labels, indent=2))


def gmail_modify(args):
    service = build_service("gmail", "v1", GMAIL_TOKEN_PATH)
    body = {}
    if args.add_labels:
        body["addLabelIds"] = args.add_labels.split(",")
    if args.remove_labels:
        body["removeLabelIds"] = args.remove_labels.split(",")
    result = service.users().messages().modify(userId="me", id=args.message_id, body=body).execute()
    print(json.dumps({"id": result["id"], "labels": result.get("labelIds", [])}, indent=2))


# =========================================================================
# Calendar
# =========================================================================

def calendar_list(args):
    service = build_service("calendar", "v3")
    now = datetime.now(timezone.utc)
    time_min = args.start or now.isoformat()
    time_max = args.end or (now + timedelta(days=7)).isoformat()

    # Ensure timezone info
    for val in [time_min, time_max]:
        if "T" in val and "Z" not in val and "+" not in val and "-" not in val[11:]:
            val += "Z"

    results = service.events().list(
        calendarId=args.calendar, timeMin=time_min, timeMax=time_max,
        maxResults=args.max, singleEvents=True, orderBy="startTime",
    ).execute()

    events = []
    for e in results.get("items", []):
        events.append({
            "id": e["id"],
            "summary": e.get("summary", "(no title)"),
            "start": e.get("start", {}).get("dateTime", e.get("start", {}).get("date", "")),
            "end": e.get("end", {}).get("dateTime", e.get("end", {}).get("date", "")),
            "location": e.get("location", ""),
            "description": e.get("description", ""),
            "status": e.get("status", ""),
            "htmlLink": e.get("htmlLink", ""),
        })
    print(json.dumps(events, indent=2, ensure_ascii=False))


def calendar_create(args):
    service = build_service("calendar", "v3")
    event = {
        "summary": args.summary,
        "start": {"dateTime": args.start},
        "end": {"dateTime": args.end},
    }
    if args.location:
        event["location"] = args.location
    if args.description:
        event["description"] = args.description
    if args.attendees:
        event["attendees"] = [{"email": e.strip()} for e in args.attendees.split(",")]

    result = service.events().insert(calendarId=args.calendar, body=event).execute()
    print(json.dumps({
        "status": "created",
        "id": result["id"],
        "summary": result.get("summary", ""),
        "htmlLink": result.get("htmlLink", ""),
    }, indent=2))


def calendar_delete(args):
    service = build_service("calendar", "v3")
    service.events().delete(calendarId=args.calendar, eventId=args.event_id).execute()
    print(json.dumps({"status": "deleted", "eventId": args.event_id}))


# =========================================================================
# Drive
# =========================================================================

def drive_search(args):
    service = build_service("drive", "v3")
    query = f"fullText contains '{args.query}'" if not args.raw_query else args.query
    if hasattr(args, 'parent') and args.parent:
        query += f" and '{args.parent}' in parents"
    results = service.files().list(
        q=query, pageSize=args.max, fields="files(id, name, mimeType, modifiedTime, webViewLink)",
    ).execute()
    files = results.get("files", [])
    print(json.dumps(files, indent=2, ensure_ascii=False))


def parse_xlsx(file_path, max_rows=500):
    """Parse Excel file to JSON."""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c is not None else "" for c in row]
            elif i <= max_rows:
                rows.append([str(c) if c is not None else "" for c in row])
        sheets.append({"name": sheet_name, "headers": headers or [], "rows": rows, "row_count": len(rows)})
    wb.close()
    return json.dumps({"type": "spreadsheet", "sheets": sheets}, indent=2, ensure_ascii=False)


def parse_docx(file_path):
    """Parse Word document to text."""
    import docx
    doc = docx.Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def parse_pdf(file_path):
    """Parse PDF to text."""
    import fitz  # pymupdf
    doc = fitz.open(file_path)
    pages = []
    for page in doc:
        text = page.get_text()
        if text.strip():
            pages.append(text)
    doc.close()
    return "\n\n--- Page Break ---\n\n".join(pages)


BINARY_HANDLERS = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": (".xlsx", parse_xlsx),
    "application/vnd.ms-excel": (".xls", parse_xlsx),
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": (".docx", parse_docx),
    "application/pdf": (".pdf", parse_pdf),
}


def drive_get(args):
    """Download and print file content by ID. Handles Google Docs/Sheets export and binary files."""
    import io
    import tempfile
    service = build_service("drive", "v3")

    # Get file metadata first
    meta = service.files().get(fileId=args.file_id, fields="name,mimeType").execute()
    mime = meta.get("mimeType", "")
    name = meta.get("name", "unknown")

    # Google native formats: export as plain text
    export_map = {
        "application/vnd.google-apps.document": "text/plain",
        "application/vnd.google-apps.spreadsheet": "text/csv",
        "application/vnd.google-apps.presentation": "text/plain",
    }

    if mime in export_map:
        content = service.files().export(fileId=args.file_id, mimeType=export_map[mime]).execute()
        if isinstance(content, bytes):
            content = content.decode("utf-8", errors="replace")
        print(content)
        return

    # Binary formats with dedicated parsers
    if mime in BINARY_HANDLERS:
        ext, handler = BINARY_HANDLERS[mime]
        from googleapiclient.http import MediaIoBaseDownload
        request = service.files().get_media(fileId=args.file_id)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(buf.getvalue())
            tmp_path = tmp.name
        try:
            result = handler(tmp_path)
            print(result)
        finally:
            os.unlink(tmp_path)
        return

    # Other binary/text files: download and try to decode as text
    from googleapiclient.http import MediaIoBaseDownload
    request = service.files().get_media(fileId=args.file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    raw = buf.getvalue()
    try:
        print(raw.decode("utf-8"))
    except UnicodeDecodeError:
        print(f"[Binary file: {name} ({mime}, {len(raw)} bytes). Cannot display as text.]")


# =========================================================================
# Contacts
# =========================================================================

def contacts_list(args):
    service = build_service("people", "v1")
    results = service.people().connections().list(
        resourceName="people/me",
        pageSize=args.max,
        personFields="names,emailAddresses,phoneNumbers",
    ).execute()
    contacts = []
    for person in results.get("connections", []):
        names = person.get("names", [{}])
        emails = person.get("emailAddresses", [])
        phones = person.get("phoneNumbers", [])
        contacts.append({
            "name": names[0].get("displayName", "") if names else "",
            "emails": [e.get("value", "") for e in emails],
            "phones": [p.get("value", "") for p in phones],
        })
    print(json.dumps(contacts, indent=2, ensure_ascii=False))


# =========================================================================
# Sheets
# =========================================================================

def sheets_get(args):
    service = build_service("sheets", "v4")
    result = service.spreadsheets().values().get(
        spreadsheetId=args.sheet_id, range=args.range,
    ).execute()
    print(json.dumps(result.get("values", []), indent=2, ensure_ascii=False))


def sheets_update(args):
    service = build_service("sheets", "v4")
    values = json.loads(args.values)
    body = {"values": values}
    result = service.spreadsheets().values().update(
        spreadsheetId=args.sheet_id, range=args.range,
        valueInputOption="USER_ENTERED", body=body,
    ).execute()
    print(json.dumps({"updatedCells": result.get("updatedCells", 0), "updatedRange": result.get("updatedRange", "")}, indent=2))


def sheets_append(args):
    service = build_service("sheets", "v4")
    values = json.loads(args.values)
    body = {"values": values}
    result = service.spreadsheets().values().append(
        spreadsheetId=args.sheet_id, range=args.range,
        valueInputOption="USER_ENTERED", insertDataOption="INSERT_ROWS", body=body,
    ).execute()
    print(json.dumps({"updatedCells": result.get("updates", {}).get("updatedCells", 0)}, indent=2))


# =========================================================================
# Docs
# =========================================================================

def docs_get(args):
    service = build_service("docs", "v1")
    doc = service.documents().get(documentId=args.doc_id).execute()
    # Extract plain text from the document structure
    text_parts = []
    for element in doc.get("body", {}).get("content", []):
        paragraph = element.get("paragraph", {})
        for pe in paragraph.get("elements", []):
            text_run = pe.get("textRun", {})
            if text_run.get("content"):
                text_parts.append(text_run["content"])
    result = {
        "title": doc.get("title", ""),
        "documentId": doc.get("documentId", ""),
        "body": "".join(text_parts),
    }
    print(json.dumps(result, indent=2, ensure_ascii=False))


# =========================================================================
# CLI parser
# =========================================================================

def main():
    parser = argparse.ArgumentParser(description="Google Workspace API for Hermes Agent")
    sub = parser.add_subparsers(dest="service", required=True)

    # --- Gmail ---
    gmail = sub.add_parser("gmail")
    gmail_sub = gmail.add_subparsers(dest="action", required=True)

    p = gmail_sub.add_parser("search")
    p.add_argument("query", help="Gmail search query (e.g. 'is:unread')")
    p.add_argument("--max", type=int, default=10)
    p.set_defaults(func=gmail_search)

    p = gmail_sub.add_parser("get")
    p.add_argument("message_id")
    p.set_defaults(func=gmail_get)

    p = gmail_sub.add_parser("send")
    p.add_argument("--to", required=True)
    p.add_argument("--subject", required=True)
    p.add_argument("--body", required=True)
    p.add_argument("--cc", default="")
    p.add_argument("--html", action="store_true", help="Send body as HTML")
    p.add_argument("--thread-id", default="", help="Thread ID for threading")
    p.set_defaults(func=gmail_send)

    p = gmail_sub.add_parser("reply")
    p.add_argument("message_id", help="Message ID to reply to")
    p.add_argument("--body", required=True)
    p.set_defaults(func=gmail_reply)

    p = gmail_sub.add_parser("labels")
    p.set_defaults(func=gmail_labels)

    p = gmail_sub.add_parser("modify")
    p.add_argument("message_id")
    p.add_argument("--add-labels", default="", help="Comma-separated label IDs to add")
    p.add_argument("--remove-labels", default="", help="Comma-separated label IDs to remove")
    p.set_defaults(func=gmail_modify)

    # --- Calendar ---
    cal = sub.add_parser("calendar")
    cal_sub = cal.add_subparsers(dest="action", required=True)

    p = cal_sub.add_parser("list")
    p.add_argument("--start", default="", help="Start time (ISO 8601)")
    p.add_argument("--end", default="", help="End time (ISO 8601)")
    p.add_argument("--max", type=int, default=25)
    p.add_argument("--calendar", default="primary")
    p.set_defaults(func=calendar_list)

    p = cal_sub.add_parser("create")
    p.add_argument("--summary", required=True)
    p.add_argument("--start", required=True, help="Start (ISO 8601 with timezone)")
    p.add_argument("--end", required=True, help="End (ISO 8601 with timezone)")
    p.add_argument("--location", default="")
    p.add_argument("--description", default="")
    p.add_argument("--attendees", default="", help="Comma-separated email addresses")
    p.add_argument("--calendar", default="primary")
    p.set_defaults(func=calendar_create)

    p = cal_sub.add_parser("delete")
    p.add_argument("event_id")
    p.add_argument("--calendar", default="primary")
    p.set_defaults(func=calendar_delete)

    # --- Drive ---
    drv = sub.add_parser("drive")
    drv_sub = drv.add_subparsers(dest="action", required=True)

    p = drv_sub.add_parser("search")
    p.add_argument("query")
    p.add_argument("--max", type=int, default=10)
    p.add_argument("--raw-query", action="store_true", help="Use query as raw Drive API query")
    p.add_argument("--parent", help="Filter by parent folder ID")
    p.set_defaults(func=drive_search)

    p = drv_sub.add_parser("get")
    p.add_argument("file_id", help="Google Drive file ID")
    p.set_defaults(func=drive_get)

    # --- Contacts ---
    con = sub.add_parser("contacts")
    con_sub = con.add_subparsers(dest="action", required=True)

    p = con_sub.add_parser("list")
    p.add_argument("--max", type=int, default=50)
    p.set_defaults(func=contacts_list)

    # --- Sheets ---
    sh = sub.add_parser("sheets")
    sh_sub = sh.add_subparsers(dest="action", required=True)

    p = sh_sub.add_parser("get")
    p.add_argument("sheet_id")
    p.add_argument("range")
    p.set_defaults(func=sheets_get)

    p = sh_sub.add_parser("update")
    p.add_argument("sheet_id")
    p.add_argument("range")
    p.add_argument("--values", required=True, help="JSON array of arrays")
    p.set_defaults(func=sheets_update)

    p = sh_sub.add_parser("append")
    p.add_argument("sheet_id")
    p.add_argument("range")
    p.add_argument("--values", required=True, help="JSON array of arrays")
    p.set_defaults(func=sheets_append)

    # --- Docs ---
    docs = sub.add_parser("docs")
    docs_sub = docs.add_subparsers(dest="action", required=True)

    p = docs_sub.add_parser("get")
    p.add_argument("doc_id")
    p.set_defaults(func=docs_get)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
